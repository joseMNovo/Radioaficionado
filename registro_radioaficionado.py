
from tkinter import *
from tkinter import OptionMenu, StringVar, ttk, filedialog
from tkinter.filedialog import askopenfile, askopenfilename
from tkinter import messagebox
from tkinter.constants import CENTER, NO, RIGHT, TOP, Y
from openpyxl import Workbook
from openpyxl import load_workbook
import os
import time
#-----------------------------------------------------------------------------------------



# Funciones ------------------------------------------------------------------------------
def buscar():
    ws = wb.active
    interior_call = call_entry.get().upper()

    listado_excel = []
    for row in ws.iter_rows(values_only = True):
        listado_excel.append((row[1], row[2], row[3]))

    for elements in listado_excel:
        for element in elements:
            indice = elements.index(element)

            if interior_call == element:
                lupa_label = Label(text=f"Call registrado en banda {elements[indice+1]} y modo {elements[indice+2]}", font = ("Arial", 15), fg = "green")
                lupa_label.place(x = 120, y = 70)
                ventana.after(5000, (lupa_label.place_forget))



def guardar(cajas):
    ws = wb.active
    ws.append(cajas)
    wb.save (filename = "radioaficionado.xlsx")


def salir():
    response = messagebox.askyesno(title = "Advertencia", message = "¿Desea salir?")
    if response:
        return ventana.destroy()


def examinar():
    #file = askopenfile(mode ='w', filetypes =[("Archivos Excel", ".xlsx .xls")])
    original_file = askopenfilename(filetypes =[("Archivos Excel", ".xlsx .xls")])
    print(type(original_file))
    #file = original_file("/")
    #print(file[2])
    #return file[2]


def comprobarArchivo():
    existe = os.path.exists("radioaficionado.xlsx") # comprueba si existe un archivo 
    if existe:
        wb = load_workbook(filename = "radioaficionado.xlsx")
        ws = wb.active
    else:
        wb = Workbook() # Creo el Excel
        ws = wb.active  # Selecciona la hoja activa
        titulo = ["ID", "CALL", "BANDA", "MODO", "FECHA"] # Agrego al principio del Excel
        ws.append(titulo) # Se guarda al final
        wb.save(filename = "radioaficionado.xlsx")

    return wb
    

def limpiar():
    call_entry.delete(0, END)
    clicked_banda.set("Seleccionar")
    clicked_modo.set("Seleccionar")


def registrar():
    ws = wb.active
    interior_call = call_entry.get().upper()
    interior_banda = clicked_banda.get()  
    interior_modo = clicked_modo.get()  
    #--------------------
    id_ingresados = ws["A"]

    lista_id = []
    for c in id_ingresados:
        lista_id.append(c.value)

    try:
        id = (int(lista_id[-1]) + 1)
    except ValueError:
        id = 1

    #-----------------------

    interior_call_banda_modo = (interior_call, interior_banda, interior_modo)
    listado_excel = []
    for row in ws.iter_rows(values_only = True):
        listado_excel.append((row[1], row[2], row[3]))

    if interior_call:
        if interior_call_banda_modo in listado_excel:
            respuesta = messagebox.askyesno(title = "Advertencia", message = "La CALL ingresada ya existe en la BANDA y MODO seleccionado. ¿Desea registrarlo?")
            if not respuesta:
                clicked_banda.set("Seleccionar")
                clicked_modo.set("Seleccionar")
            else:
    # --------------------------------------------------------------------------------------------------------------------
            # Programa principal PRIMER ELSE
                if interior_banda != "Seleccionar":
                    if interior_modo != "Seleccionar":
                        tiempo = time.gmtime()
                        fecha = f"{tiempo[2]}/{tiempo[1]}/{tiempo[0]} - {tiempo[3]}:{tiempo[4]}'{tiempo[5]}''"
                        
                        datos = id, interior_call, interior_banda, interior_modo, fecha
                        #datos_para_base = [id, interior_call, interior_banda, interior_modo,fecha]

                        # Se guarda en Excel
                        guardar(datos)
                        
                        # Se guarda en Treeview
                        if id % 2 == 0:
                            treeview.insert(parent = "", index = "end", text = "", values = datos, tags = ("evenrow"))
                        else:
                            treeview.insert(parent = "", index = "end", text = "", values = datos, tags = ("oddrow"))

                        """ et_confirmacion_registro = Label(text="Registro exitoso", font = ("Arial", 20), fg = "green")
                        et_confirmacion_registro.place(x = 95, y = 205)
                        ventana.after(1500, (et_confirmacion_registro.place_forget)) """
                    else: 
                        messagebox.showwarning(title = "Advertencia", message = "Seleccione el MODO")

                else:
                    messagebox.showwarning(title = "Advertencia", message = "Seleccione la BANDA")
    # --------------------------------------------------------------------------------------------------------------------
    # --------------------------------------------------------------------------------------------------------------------
        else:
        # Programa principal SEGUNDO ELSE
            if interior_banda != "Seleccionar":
                if interior_modo != "Seleccionar":
                    tiempo = time.gmtime()
                    fecha = f"{tiempo[2]}/{tiempo[1]}/{tiempo[0]} - {tiempo[3]}:{tiempo[4]}:{tiempo[5]}"
                    
                    datos = id, interior_call, interior_banda, interior_modo, fecha
                    #datos_para_base = [id, interior_call, interior_banda, interior_modo,fecha]

                    # Se guarda en Excel
                    guardar(datos)
                    
                    # Se guarda en Treeview
                    if id % 2 == 0:
                        treeview.insert(parent = "", index = "end", text = "", values = datos, tags = ("evenrow"))
                    else:
                        treeview.insert(parent = "", index = "end", text = "", values = datos, tags = ("oddrow"))

                    """ et_confirmacion_registro = Label(text="Registro exitoso", font = ("Arial", 20), fg = "green")
                    et_confirmacion_registro.place(x = 95, y = 205)
                    ventana.after(1500, (et_confirmacion_registro.place_forget)) """
                else: 
                    messagebox.showwarning(title = "Advertencia", message = "Seleccione el MODO")

            else:
                messagebox.showwarning(title = "Advertencia", message = "Seleccione la BANDA")
    # --------------------------------------------------------------------------------------------------------------------
        

    else:
        messagebox.showwarning(title = "Advertencia", message = "Ingrese la CALL")


#-----------------------------------------------------------------------------------------
wb = comprobarArchivo()

ventana = Tk()
ventana.title("Radioaficionado")
ventana.geometry("710x600")
ventana.resizable(False, False)


# Frames ----------------------------------------------------------------
titulo_label = Label(text = "José Manuel L. Novo - LU4FYF", font= ("Arial 25 underline"))
titulo_label.pack(side=TOP, pady=20)

tree_frame = Frame(ventana)
tree_frame.pack(side=BOTTOM, ipady=5)

campos_frame = LabelFrame(ventana, text = "Campos")
campos_frame.pack(side = LEFT, ipadx=5, padx = 10)

botones_frame = LabelFrame(ventana, text = "Comandos")
botones_frame.pack(side = RIGHT, ipadx=5, padx = 10)



    
# Scrollbar ----------------------------------------------------------------
tree_scroll = Scrollbar(tree_frame)
tree_scroll.pack(side = RIGHT, fill = Y)

# Treeview ----------------------------------------------------------------

style = ttk.Style()
style.theme_use("clam")

# Configure the Treeview Colors
style.configure("Treeview", 
    background = "#D3D3D3",
    foreground = "black",
    rowheight = 25,
    fieldbackground = "#D3D3D3")

# Change Selected color
style.map("Treeview", background = [("selected", "#347083")])

treeview = ttk.Treeview(tree_frame, yscrollcommand = tree_scroll.set, selectmode="extended")
treeview.pack()

tree_scroll.config(command=treeview.yview)

# Define our Columns
treeview["columns"] = ("ID", "CALL", "BANDA", "MODO", "FECHA")

# Format Our Columns
treeview.column("#0", width = 0, stretch = NO)
treeview.column("ID", anchor = CENTER, width = 100)
treeview.column("CALL", anchor = CENTER, width = 140)
treeview.column("BANDA", anchor = CENTER, width = 140)
treeview.column("MODO", anchor = CENTER, width = 140)
treeview.column("FECHA", anchor = CENTER, width = 140)

# Create Headings
treeview.heading("#0", text = "", anchor = CENTER)
treeview.heading("ID", text = "ID", anchor = CENTER)
treeview.heading("CALL", text = "CALL", anchor = CENTER)
treeview.heading("BANDA", text = "BANDA", anchor = CENTER)
treeview.heading("MODO", text = "MODO", anchor = CENTER)
treeview.heading("FECHA", text = "FECHA", anchor = CENTER)


# Create Striped Row Tags
treeview.tag_configure("oddrow", background="white")
treeview.tag_configure("evenrow", background="lightblue")

# Add our data to the screen

# Etiquetas ----------------------------------------------------------------
letra_grande = ("Verdana", 15)

call_label = Label(campos_frame, text="CALL", font = letra_grande)
call_label.grid(row = 0, column = 0, padx = 10, pady = 10)

banda_label = Label(campos_frame, text="BANDA", font = letra_grande)
banda_label.grid(row = 1, column = 0, padx = 10, pady = 10)

modo_label = Label(campos_frame, text="MODO", font = letra_grande)
modo_label.grid(row = 2, column = 0, padx = 10, pady = 10)



#ventana.geometry("710x600")


# Cajas ----------------------------------------------------------------
call_entry = Entry(campos_frame, font = letra_grande, width=10)
call_entry.grid(row = 0, column = 1)

# Lista (bandas) ----------------------------------------------------------------

clicked_banda = StringVar()
clicked_banda.set("Seleccionar")

clicked_modo = StringVar()
clicked_modo.set("Seleccionar")

lista_banda = OptionMenu(campos_frame, clicked_banda, "Seleccionar", "2","6","10","12","15","17","20","30","40","80","160","DV")
lista_banda.config(bg = "lightgray")
lista_banda["menu"].config(bg = "lightgray")
lista_banda.grid(row=1, column=1, padx = 10, pady = 10, ipady=6)

lista_modo = OptionMenu(campos_frame, clicked_modo, "Seleccionar", "SSB", "FT8", "FT4")
lista_modo.config(bg = "lightgray")
lista_modo["menu"].config(bg = "lightgray")
lista_modo.grid(row=2, column=1, padx = 10, pady = 10, ipady=6)


# Botones ----------------------------------------------------------------

style.configure("Registrar.TButton", font = ("Arial", 15))
style.map("Registrar.TButton", foreground=[('pressed', 'black'), ('active', 'blue')])

style.configure("Limpiar.TButton", font = ("Arial", 15))
style.map("Limpiar.TButton", foreground=[('pressed', 'black'), ('active', 'black')], background = [('active', 'lightgreen')])

style.configure("Salir.TButton", font = ("Arial", 15))
style.map("Salir.TButton", foreground=[('pressed', 'black'), ('active', 'black')], background = [('active', 'lightblue')])

style.configure("Examinar.TButton", font = ("Arial", 15))
style.map("Examinar.TButton", foreground=[('pressed', 'black'), ('active', 'black')], background = [('active', 'violet')])

registrar_button = ttk.Button(botones_frame, text = "Registrar", style = "Limpiar.TButton", width=29, command=registrar)
registrar_button.grid (row=2, column=0, padx=10, pady=5, ipady=7)

limpiar_button = ttk.Button(botones_frame, text = "Limpiar", style = "Registrar.TButton", width=29, command=limpiar)
limpiar_button.grid (row=3, column=0, padx=10, pady=5)

examinar_button = ttk.Button(botones_frame, text = "Examinar...", style = "Examinar.TButton", width=29, command= examinar)
examinar_button.grid (row=4, column=0, padx=10, pady=5)

salir_button = ttk.Button(botones_frame, text = "Salir", style = "Salir.TButton", width=10, command= salir)
salir_button.grid (row=5, column=0, padx=10, pady=5)




#lupa_img = PhotoImage(file='imagen/lupa.png', width=15, height=15)

""" buscar_button = ttk.Button(campos_frame, image=lupa_img, command=buscar)
buscar_button.grid (row=0, column=2) """


ventana.mainloop()
